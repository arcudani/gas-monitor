"""Test export (RF-05): le funzioni pure producono file validi (offline, con
fixture) e la pagina renderizza sui dati reali (se SUPABASE_DB_URL).
"""
from __future__ import annotations

import datetime as dt
import io
import os
import sys

import pandas as pd

import config  # noqa: F401
import export


def _fixture():
    giorni = pd.date_range("2026-08-01", "2026-08-18")
    seg = pd.DataFrame({
        "Data": giorni, "Scenario": [50 + i for i in range(len(giorni))],
        "ScenarioM": [52.0] * len(giorni), "Punteggi": [{"stoccaggi": 80, "meteo": 10, "lng": 20, "geopolitica": 60}] * len(giorni),
        "Prezzo": [60 - i * 0.3 for i in range(len(giorni))],
        "TBreve": [61.0] * len(giorni), "TMedio": [63.0] * len(giorni), "TLungo": [58.0] * len(giorni),
        "ScB": [-1.5] * len(giorni), "ScM": [-6.0] * len(giorni), "PendL": [0.1] * len(giorni),
        "PzOk": [False] * 5 + [True] * 13, "N": list(range(len(giorni))),
        "Codice": ["attesa"] * 5 + ["opportunita"] * 13, "Testo": ["testo di prova"] * len(giorni),
    })
    var = pd.DataFrame({
        "Data": list(giorni) * 2,
        "Metrica": ["riempimento_pct"] * len(giorni) + ["gpr"] * len(giorni),
        "Valore": [80.0] * len(giorni) + [120.0] * len(giorni),
    })
    return seg, var


def test_excel_valido() -> None:
    from openpyxl import load_workbook
    seg, var = _fixture()
    b = export.excel_serie(seg, var, dt.date(2026, 8, 1), dt.date(2026, 8, 18))
    wb = load_workbook(io.BytesIO(b))
    assert wb.sheetnames == ["Segnale", "Variabili", "Note"]
    ws = wb["Segnale"]
    assert ws.max_row == len(seg) + 1 and ws["J7"].value == "Opportunità di prezzo"
    assert wb["Variabili"].max_row == len(seg) + 1
    assert any(export.DISCLAIMER in str(c.value) for r in wb["Note"].iter_rows() for c in r if c.value)
    print(f"OK - excel valido ({len(b)} byte, 3 fogli)")


def test_pdf_valido() -> None:
    seg, _ = _fixture()
    u = seg.iloc[-1]
    sit = dict(data=u["Data"], scenario=float(u["Scenario"]), scenario_m=52.0,
               punteggi=u["Punteggi"], prezzo=float(u["Prezzo"]), d_prezzo=-0.3,
               sc_b=-1.5, sc_m=-6.0, pend_l=0.1, pz_ok=True, codice="opportunita",
               n=13, testo="Opportunità di prezzo: testo di prova.")
    b = export.pdf_snapshot(sit, seg, {"scenario_soglia": 60, "scenario_media_gg": 7},
                            export.Path(__file__).parent / "assets" / "Logo_Bros_Consulenza_460.png")
    assert b[:5] == b"%PDF-" and len(b) > 3000
    print(f"OK - pdf valido ({len(b)} byte)")


def test_pagina_export() -> None:
    from streamlit.testing.v1 import AppTest
    at = AppTest.from_file("moduli/export.py", default_timeout=90)
    at.run()
    assert not at.exception, [str(e.value) for e in at.exception]
    sub = " ".join(h.value for h in at.subheader)
    assert "Snapshot PDF" in sub and "Serie storiche Excel" in sub
    print("OK - pagina Export renderizza con i due download")


if __name__ == "__main__":
    try:
        test_excel_valido()
        test_pdf_valido()
        if os.environ.get("SUPABASE_DB_URL", "").strip():
            test_pagina_export()
        else:
            print("[SKIP] pagina Export: SUPABASE_DB_URL assente")
    except AssertionError as e:
        print(f"FAIL - {e}")
        sys.exit(1)
    print("Tutti i test OK")
    sys.exit(0)
