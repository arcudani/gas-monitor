"""
Export (RF-05) — funzioni PURE che producono bytes, senza Streamlit:
  - excel_serie(df_segnale, df_variabili, dal, al) -> bytes .xlsx
  - pdf_snapshot(situazione, df_segnale, cfg, logo_path) -> bytes .pdf
Testabili offline; la pagina Export le cabla con st.download_button.

Veste Bros: rosso #C00000 per titoli/righe di intestazione, blu #0F6FA8 per
l'accento del Gas Monitor, Inter non embedded (Helvetica nel PDF).
Disclaimer obbligatorio in entrambi i formati (RNF).
"""
from __future__ import annotations

import datetime as dt
import io
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (Image, Paragraph, SimpleDocTemplate, Spacer,
                                Table, TableStyle)

ROSSO = "#C00000"
BLU = "#0F6FA8"
DISCLAIMER = ("Le informazioni hanno scopo informativo e di supporto "
              "all'approvvigionamento; non costituiscono consulenza finanziaria. "
              "La decisione resta dell'utente.")
NOME_LIVELLO = {
    "attesa": "Attesa", "chiusa": "Finestra chiusa", "monitorare": "Monitorare",
    "opportunita": "Opportunità di prezzo", "minimo": "Minimo di periodo",
    "prime": "Prime condizioni",
    "iniziale": "Segnale iniziale", "fixing": "Segnale di fixing",
    "trend": "Trend consolidato",
}


def _it(v, dec: int = 2) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "n.d."
    return f"{float(v):,.{dec}f}".replace(",", "X").replace(".", ",").replace("X", ".")


# Contesto di commodity per i documenti (g008): titolo, etichette prezzo,
# variabili dello scenario e fonti. Default = gas storico, cosi' le chiamate
# esistenti (e i test offline) restano valide.
CTX_GAS = {
    "titolo": "Gas Market Monitor", "prezzo_label": "MGP-GAS", "prezzo_udm": "€/MWh",
    "prezzo_nome": "PREZZO GAS MGP-GAS",
    "variabili": [("stoccaggi", "Stoccaggi"), ("meteo", "Meteo"), ("lng", "LNG"), ("geopolitica", "Geo")],
    "fonti": ["Prezzo: GME — MGP-GAS (mercato del giorno prima gas), €/MWh",
              "Stoccaggi: GIE AGSI+ (Italia)", "LNG: GIE ALSI (Italia)",
              "Meteo: Open-Meteo / ERA5, paniere città pesato per consumo gas",
              "Geopolitica: GPR daily — Caldara & Iacoviello"],
    "fonti_breve": "GME MGP-GAS · GIE AGSI+/ALSI · Open-Meteo ERA5 · GPR Caldara-Iacoviello",
}


def contesto(info: dict, variabili: list[dict]) -> dict:
    """Costruisce il contesto documenti dall'anagrafica (commodity.info/variabili)."""
    if info.get("commodity", "gas") == "gas" and not variabili:
        return CTX_GAS
    nome = info["nome"]
    titolo = "Gas Market Monitor" if info["commodity"] == "gas" else f"Market Monitor — {nome}"
    return {
        "titolo": titolo, "prezzo_label": info["prezzo_label"], "prezzo_udm": info["prezzo_udm"],
        "prezzo_nome": f"PREZZO {nome.upper()} {info['prezzo_label']}",
        "variabili": [(v["variabile"], v["etichetta"].split(" (")[0]) for v in variabili],
        "fonti": [f"Prezzo: {info['prezzo_fonte']}, {info['prezzo_udm']}"]
                 + [f"{v['etichetta']}: {v['fonte']}" for v in variabili],
        "fonti_breve": " · ".join([info["prezzo_fonte"].split(",")[0]]
                                  + sorted({v["fonte"] for v in variabili})),
    }


# =============================================================================
# EXCEL
# =============================================================================

def excel_serie(df_seg: pd.DataFrame, df_var: pd.DataFrame,
                dal: dt.date, al: dt.date, ctx: dict | None = None) -> bytes:
    """Workbook con 3 fogli: Segnale (giorno per giorno), Variabili (serie
    pivotate per giorno), Note (fonti, periodo, disclaimer).
    df_seg colonne: Data, Scenario, ScenarioM, Prezzo, TBreve, TMedio, TLungo,
      ScB, ScM, Codice, N, Testo.  df_var: Data, Metrica, Valore.
    ctx: contesto commodity (vedi contesto()); default gas."""
    ctx = ctx or CTX_GAS
    wb = Workbook()
    h_font = Font(bold=True, color="FFFFFF")
    h_fill = PatternFill("solid", fgColor="C00000")

    def _intesta(ws, cols):
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font, cell.fill = h_font, h_fill
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    # --- Segnale ---
    ws = wb.active
    ws.title = "Segnale"
    cols = ["Data", "Scenario", "Scenario media", f"Prezzo {ctx['prezzo_udm']}", "Trend breve",
            "Trend medio", "Trend lungo", "vs breve %", "vs medio %",
            "Segnale", "Giorni consecutivi", "Testo"]
    _intesta(ws, cols)
    for _, r in df_seg.iterrows():
        ws.append([
            pd.Timestamp(r["Data"]).date(),
            None if pd.isna(r["Scenario"]) else round(float(r["Scenario"]), 1),
            None if pd.isna(r.get("ScenarioM")) else round(float(r["ScenarioM"]), 1),
            None if pd.isna(r["Prezzo"]) else round(float(r["Prezzo"]), 4),
            None if pd.isna(r["TBreve"]) else round(float(r["TBreve"]), 4),
            None if pd.isna(r["TMedio"]) else round(float(r["TMedio"]), 4),
            None if pd.isna(r["TLungo"]) else round(float(r["TLungo"]), 4),
            None if pd.isna(r["ScB"]) else round(float(r["ScB"]), 2),
            None if pd.isna(r["ScM"]) else round(float(r["ScM"]), 2),
            NOME_LIVELLO.get(r["Codice"], r["Codice"]),
            int(r["N"]) if not pd.isna(r["N"]) else 0,
            r["Testo"],
        ])
    for i, w in enumerate([12, 10, 14, 13, 12, 12, 12, 11, 11, 22, 10, 80], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        row[0].number_format = "DD/MM/YYYY"

    # --- Variabili (pivot per data) ---
    ws2 = wb.create_sheet("Variabili")
    piv = (df_var.pivot_table(index="Data", columns="Metrica", values="Valore",
                              aggfunc="first").sort_index()
           if not df_var.empty else pd.DataFrame())
    nomi = {"t_media_paniere": "T media paniere °C",
            "riempimento_pct": "Stoccaggi IT %",
            "iniezione": "Iniezione GWh/g", "erogazione": "Erogazione GWh/g",
            "lng_sendout": "LNG send-out GWh/g", "lng_giacenza": "LNG giacenza GWh",
            "gpr": "GPR daily"}
    colv = ["Data"] + [nomi.get(c, c) for c in piv.columns]
    _intesta(ws2, colv)
    for d, r in piv.iterrows():
        ws2.append([pd.Timestamp(d).date()] +
                   [None if pd.isna(v) else round(float(v), 3) for v in r.values])
    ws2.column_dimensions["A"].width = 12
    for i in range(2, len(colv) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 18
    for row in ws2.iter_rows(min_row=2, min_col=1, max_col=1):
        row[0].number_format = "DD/MM/YYYY"

    # --- Note ---
    ws3 = wb.create_sheet("Note")
    righe = [
        [f"{ctx['titolo']} — Bros Consulenza s.r.l."],
        [f"Periodo: {dal.strftime('%d/%m/%Y')} – {al.strftime('%d/%m/%Y')}"],
        [f"Generato il {dt.datetime.now().strftime('%d/%m/%Y %H:%M')}"],
        [],
        ["Fonti"],
        *[[f] for f in ctx["fonti"]],
        [],
        ["Scenario 0–100: fondamentali orientati a favore dell'acquirente (rango percentile 2020–oggi)."],
        ["Segnale: scenario (media mobile) ≥ soglia e prezzo in flessione vs trend, per n giorni consecutivi; "
         "opportunità di prezzo = prezzo in netta flessione sul trend medio con scenario non sfavorevole."],
        [],
        [DISCLAIMER],
    ]
    for r in righe:
        ws3.append(r)
    ws3["A1"].font = Font(bold=True, color="C00000", size=13)
    ws3["A5"].font = Font(bold=True)
    ws3.column_dimensions["A"].width = 120

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
# PDF
# =============================================================================

def pdf_snapshot(sit: dict, df_seg: pd.DataFrame, cfg: dict,
                 logo_path: Path | None = None, ctx: dict | None = None) -> bytes:
    """Snapshot A4 della situazione: intestazione Bros, pannelli scenario/
    prezzo/segnale, tabella ultimi 14 giorni, fonti e disclaimer.
    sit: {data, scenario, scenario_m, punteggi{}, prezzo, d_prezzo, sc_b, sc_m,
          pend_l, codice, n, testo}. ctx: contesto commodity (default gas)."""
    ctx = ctx or CTX_GAS
    udm = ctx["prezzo_udm"]
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=18 * mm, rightMargin=18 * mm,
                            topMargin=14 * mm, bottomMargin=14 * mm,
                            title=f"{ctx['titolo']} — snapshot",
                            author="Bros Consulenza s.r.l.")
    st_t = ParagraphStyle("t", fontName="Helvetica-Bold", fontSize=17,
                          textColor=colors.HexColor(ROSSO), spaceAfter=2)
    st_s = ParagraphStyle("s", fontName="Helvetica", fontSize=9.5,
                          textColor=colors.HexColor("#64748b"), spaceAfter=8)
    st_h = ParagraphStyle("h", fontName="Helvetica-Bold", fontSize=8,
                          textColor=colors.HexColor("#64748b"))
    st_big = ParagraphStyle("b", fontName="Helvetica-Bold", fontSize=30,
                            textColor=colors.HexColor(BLU), leading=34)
    st_big_k = ParagraphStyle("bk", parent=st_big, textColor=colors.HexColor("#0f172a"))
    st_p = ParagraphStyle("p", fontName="Helvetica", fontSize=9.5, leading=13)
    st_small = ParagraphStyle("sm", fontName="Helvetica", fontSize=8,
                              textColor=colors.HexColor("#64748b"), leading=11)
    st_seg = ParagraphStyle("seg", fontName="Helvetica-Bold", fontSize=15,
                            textColor=colors.HexColor(BLU), leading=18)

    el = []
    # intestazione
    testata = [[Paragraph(ctx["titolo"], st_t),
                Image(str(logo_path), width=28 * mm, height=28 * mm * 0.62)
                if logo_path and Path(logo_path).exists() else ""]]
    t = Table(testata, colWidths=[140 * mm, 34 * mm])
    t.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                           ("ALIGN", (1, 0), (1, 0), "RIGHT")]))
    el += [t, Paragraph(
        f"Situazione al {sit['data'].strftime('%d/%m/%Y')} · "
        f"Bros Consulenza s.r.l. · generato il {dt.datetime.now().strftime('%d/%m/%Y %H:%M')}",
        st_s)]
    el.append(Table([[""]], colWidths=[174 * mm], rowHeights=[2],
                    style=[("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(ROSSO))]))
    el.append(Spacer(1, 6))

    # pannelli scenario | prezzo
    p = sit.get("punteggi") or {}
    scen_fascia = ("favorevole" if sit["scenario_m"] >= float(cfg.get("scenario_soglia", 60))
                   else "neutro" if sit["scenario_m"] >= 40 else "sfavorevole")
    sx = [Paragraph("SCENARIO DI APPROVVIGIONAMENTO", st_h),
          Paragraph(f"{sit['scenario']:.0f} <font size=10 color='#475569'>/100 · {scen_fascia}</font>", st_big),
          Paragraph(f"media {int(cfg.get('scenario_media_gg', 7))} gg: {sit['scenario_m']:.0f} "
                    f"(soglia {float(cfg.get('scenario_soglia', 60)):.0f})", st_small),
          Paragraph(" · ".join(f"{n} {p[k]:.0f}" if p.get(k) is not None else f"{n} n.d."
                               for k, n in ctx["variabili"]), st_small)]
    ver = "prezzo in flessione" if sit.get("pz_ok") else "prezzo non in flessione"
    dx = [Paragraph(ctx["prezzo_nome"], st_h),
          Paragraph(f"{_it(sit['prezzo'], 2)} <font size=10 color='#475569'>{udm} · {ver}</font>", st_big_k),
          Paragraph(f"vs trend breve {_pct(sit['sc_b'])} · vs trend medio {_pct(sit['sc_m'])} · "
                    f"trend lungo {_dir(sit['pend_l'])}", st_small),
          Paragraph(f"{'▲' if sit['d_prezzo'] > 0 else '▼' if sit['d_prezzo'] < 0 else '='} "
                    f"{_it(abs(sit['d_prezzo']), 2)} {udm} vs giorno precedente", st_small)]
    tp = Table([[sx, dx]], colWidths=[87 * mm, 87 * mm])
    tp.setStyle(TableStyle([
        ("BOX", (0, 0), (0, 0), 0.6, colors.HexColor("#e7eaf0")),
        ("BOX", (1, 0), (1, 0), 0.6, colors.HexColor("#e7eaf0")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 7), ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    el += [tp, Spacer(1, 6)]

    # segnale
    nome = NOME_LIVELLO.get(sit["codice"], sit["codice"])
    n = int(sit.get("n") or 0)
    LIV = {"attesa": 0, "chiusa": 0, "monitorare": 1, "opportunita": 2, "prime": 2,
           "minimo": 3, "iniziale": 3, "fixing": 4, "trend": 5}
    liv = LIV.get(sit["codice"], 0)
    tacche = "■" * liv + "□" * (5 - liv)
    seg = Table([[[Paragraph("SEGNALE DEL GIORNO", st_h),
                   Paragraph(f"{nome} &nbsp;<font size=10 color='#64748b'>{tacche} intensità {liv}/5</font>", st_seg),
                   Paragraph(f"{n} giorni consecutivi favorevoli" if n else
                             "nessun giorno favorevole in corso", st_small),
                   Spacer(1, 3),
                   Paragraph(sit["testo"], st_p)]]], colWidths=[174 * mm])
    seg.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#e7eaf0")),
        ("LINEBEFORE", (0, 0), (0, 0), 4, colors.HexColor(BLU)),
        ("LEFTPADDING", (0, 0), (-1, -1), 10), ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    el += [seg, Spacer(1, 10)]

    # tabella ultimi 14 giorni
    el.append(Paragraph("Ultimi 14 giorni", ParagraphStyle(
        "h2", fontName="Helvetica-Bold", fontSize=11, textColor=colors.HexColor("#0f172a"),
        spaceAfter=4)))
    ultimi = df_seg.tail(14).iloc[::-1]
    righe = [["Data", "Scenario", "Media", f"Prezzo {udm}", "vs breve", "vs medio", "Segnale"]]
    for _, r in ultimi.iterrows():
        righe.append([pd.Timestamp(r["Data"]).strftime("%d/%m"),
                      f"{float(r['Scenario']):.0f}" if pd.notna(r["Scenario"]) else "–",
                      f"{float(r['ScenarioM']):.0f}" if pd.notna(r.get("ScenarioM")) else "–",
                      _it(r["Prezzo"], 2), _pct(r["ScB"]), _pct(r["ScM"]),
                      NOME_LIVELLO.get(r["Codice"], r["Codice"])])
    tt = Table(righe, colWidths=[16 * mm, 18 * mm, 16 * mm, 26 * mm, 20 * mm, 20 * mm, 58 * mm],
               repeatRows=1)
    tt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(ROSSO)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("ALIGN", (1, 1), (5, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f7f9")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e7eaf0")),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    el += [tt, Spacer(1, 10)]

    el.append(Paragraph(
        f"Fonti: {ctx['fonti_breve']}. "
        "Scenario = fondamentali orientati a favore dell'acquirente (rango percentile 2020–oggi). "
        "Segnale = scenario (media mobile) ≥ soglia e prezzo in flessione vs trend per n giorni; "
        "opportunità di prezzo = prezzo in netta flessione sul trend medio con scenario non sfavorevole.",
        st_small))
    el.append(Spacer(1, 4))
    el.append(Paragraph(DISCLAIMER, ParagraphStyle(
        "d", parent=st_small, textColor=colors.HexColor("#475569"))))
    doc.build(el)
    return buf.getvalue()


def _pct(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "n.d."
    v = float(v)
    return f"{'+' if v >= 0 else '−'}{_it(abs(v), 1)}%"


def _dir(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "n.d."
    v = float(v)
    return "in salita" if v > 0.05 else "in discesa" if v < -0.05 else "piatto"
